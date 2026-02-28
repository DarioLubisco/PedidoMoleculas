import openpyxl

old_file = r'C:\OneDrive\Farmacia Americana\Compras\Pedidos\Pedido_20251111_21Dias_200Items_Filtrado.xlsx'
new_file = r'C:\OneDrive\Farmacia Americana\Compras\Pedidos\Pedido_20260228_30Dias_200Items.xlsx'

try:
    print('--- OLD FILE ---')
    wb = openpyxl.load_workbook(old_file, data_only=True)
    ws = wb.active
    print('A2 Data Type:', ws['A2'].data_type)
    print('A2 Value:', repr(ws['A2'].value))
    print('A2 Number Format:', ws['A2'].number_format)
    print('B2 Data Type:', ws['B2'].data_type)
    print('B2 Value:', repr(ws['B2'].value))
    print('B2 Number Format:', ws['B2'].number_format)
except Exception as e:
    print('Error old file:', e)

print('\n----------------\n')

try:
    print('--- NEW FILE ---')
    wb = openpyxl.load_workbook(new_file, data_only=True)
    ws = wb.active
    print('A2 Data Type:', ws['A2'].data_type)
    print('A2 Value:', repr(ws['A2'].value))
    print('A2 Number Format:', ws['A2'].number_format)
    print('B2 Data Type:', ws['B2'].data_type)
    print('B2 Value:', repr(ws['B2'].value))
    print('B2 Number Format:', ws['B2'].number_format)
except Exception as e:
    print('Error new file:', e)
